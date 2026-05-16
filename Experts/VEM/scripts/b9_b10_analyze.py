#!/usr/bin/env python3
"""B9/B10 analysis: trades from ReportTester xlsx + EURUSD M5 bars from MT5."""
import sys
from collections import defaultdict
from datetime import datetime, timezone

import MetaTrader5 as mt5
import numpy as np
import openpyxl

TERMINAL_EXE = r"C:\Program Files\MetaTrader 5\terminal64.exe"
XLSX = (
    r"C:\Users\emili\AppData\Roaming\MetaQuotes\Terminal"
    r"\D0E8209F77C8CF37AD8BF550E51FF075\MQL5\Profiles\Tester"
    r"\ReportTester-23489.xlsx"
)
BB_PERIOD, BB_DEV = 20, 2.0
SIGNAL_SHIFT = 1
M5_SEC = 300
MAX_BARS = 100000


def load_trades(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    start = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row[0] == "Time" and row[1] == "Deal":
            start = i + 1
            break
    ins = {}
    for row in ws.iter_rows(min_row=start, values_only=True):
        if row[2] == "EURUSD" and row[4] == "in":
            ins[int(row[1])] = (row[0], row[3])
    trades = []
    for row in ws.iter_rows(min_row=start, values_only=True):
        if row[2] != "EURUSD" or row[4] != "out":
            continue
        d_in = int(row[1]) - 1
        if d_in not in ins:
            continue
        et, side = ins[d_in]
        if isinstance(et, datetime):
            entry_dt = et
        else:
            entry_dt = datetime.strptime(str(et), "%Y.%m.%d %H:%M:%S")
        # Tester timestamps match chart/server time; treat as UTC for bar lookup.
        if entry_dt.tzinfo is None:
            entry_dt = entry_dt.replace(tzinfo=timezone.utc)
        trades.append(
            {
                "entry": entry_dt,
                "side": side,
                "profit": float(row[10] or 0),
                "exit": str(row[12] or ""),
            }
        )
    wb.close()
    return trades


def load_rates():
    if not mt5.initialize(path=TERMINAL_EXE):
        raise RuntimeError(f"MT5 init failed: {mt5.last_error()}")
    if not mt5.symbol_select("EURUSD", True):
        raise RuntimeError(f"symbol_select failed: {mt5.last_error()}")
    parts = []
    for start in range(0, MAX_BARS, 50000):
        r = mt5.copy_rates_from_pos("EURUSD", mt5.TIMEFRAME_M5, start, 50000)
        if r is None or len(r) == 0:
            break
        parts.append(r)
    mt5.shutdown()
    if not parts:
        raise RuntimeError("No M5 rates from MT5")
    rates = np.concatenate(parts)
    rates = rates[np.argsort(rates["time"])]
    _, idx = np.unique(rates["time"], return_index=True)
    rates = rates[idx]
    return rates


def calc_bb(close, i, period=20, dev=2.0):
    if i < period - 1:
        return np.nan, np.nan, np.nan
    w = close[i - period + 1 : i + 1]
    mid = w.mean()
    std = w.std(ddof=0)
    return mid + dev * std, mid, mid - dev * std


def analyze(trades, rates):
    t = rates["time"].astype(np.int64)
    o, h, l, c = rates["open"], rates["high"], rates["low"], rates["close"]
    n = len(c)
    time_to_idx = {int(t[i]): i for i in range(n)}
    bb_up = np.full(n, np.nan)
    bb_lo = np.full(n, np.nan)
    for i in range(n):
        u, _, lo = calc_bb(c, i)
        bb_up[i], bb_lo[i] = u, lo

    t_min = int(t[0])
    t_max = int(t[-1])

    def signal_index(entry_dt):
        ts = int(entry_dt.timestamp()) - M5_SEC * SIGNAL_SHIFT
        if ts in time_to_idx:
            return time_to_idx[ts]
        idx = int(np.searchsorted(t, ts, side="right") - 1)
        if idx < 0 or idx >= n:
            return None
        # allow 1 bar tolerance if clock/skew
        if abs(int(t[idx]) - ts) > M5_SEC * 2:
            return None
        return idx

    def walk_count(idx, is_long, max_look=10):
        cnt = 0
        for j in range(1, max_look + 1):
            k = idx - j
            if k < 0 or np.isnan(bb_lo[k]):
                break
            if is_long:
                if c[k] < bb_lo[k]:
                    cnt += 1
                else:
                    break
            else:
                if c[k] > bb_up[k]:
                    cnt += 1
                else:
                    break
        return cnt

    def wick_pct(idx, is_long):
        rng = h[idx] - l[idx]
        if rng <= 0:
            return 0.0
        body_lo = min(o[idx], c[idx])
        body_hi = max(o[idx], c[idx])
        wick = (body_lo - l[idx]) if is_long else (h[idx] - body_hi)
        return 100.0 * wick / rng

    ok, skip = 0, 0
    for tr in trades:
        ets = int(tr["entry"].timestamp())
        if ets < t_min or ets > t_max + M5_SEC:
            skip += 1
            continue
        idx = signal_index(tr["entry"])
        if idx is None or idx < BB_PERIOD:
            skip += 1
            continue
        is_long = tr["side"] == "buy"
        tr["walk"] = walk_count(idx, is_long)
        tr["wick_pct"] = wick_pct(idx, is_long)
        ok += 1

    return ok, skip, t_min, t_max


def report(group, label):
    if not group:
        print(f"\n=== {label}: (empty) ===")
        return
    walks = np.array([t["walk"] for t in group])
    wicks = np.array([t["wick_pct"] for t in group])
    ge2 = int((walks >= 2).sum())
    print(f"\n=== {label} (N={len(group)}) ===")
    print(
        f"  BB walk:  mean={walks.mean():.2f}  median={np.median(walks):.0f}  "
        f">=2 prior closes: {ge2} ({100*ge2/len(group):.1f}%)"
    )
    print(
        f"  Wick %:   mean={wicks.mean():.1f}  median={np.median(wicks):.1f}  "
        f"p25={np.percentile(wicks,25):.1f}  p75={np.percentile(wicks,75):.1f}"
    )


def main():
    trades = load_trades(XLSX)
    rates = load_rates()
    ok, skip, t_min, t_max = analyze(trades, rates)
    bar_from = datetime.fromtimestamp(t_min, tz=timezone.utc)
    bar_to = datetime.fromtimestamp(t_max, tz=timezone.utc)
    print(f"Trades: {len(trades)}  analyzed: {ok}  skipped (no bar): {skip}")
    print(f"M5 bars: {len(rates)}  range {bar_from} .. {bar_to}")

    analyzed = [t for t in trades if "walk" in t]
    losers = [t for t in analyzed if t["profit"] < 0]
    winners = [t for t in analyzed if t["profit"] > 0]
    worst40 = sorted(losers, key=lambda x: x["profit"])[:40]
    sl_losers = [t for t in losers if "sl" in t["exit"].lower()]

    report(losers, "ALL LOSERS")
    report(winners, "ALL WINNERS")
    report(worst40, "WORST 40 LOSERS")
    report(sl_losers, "SL LOSERS (full stop)")

    if losers:
        print("\n--- Losers: walk count distribution ---")
        for w in range(0, 6):
            cnt = sum(1 for t in losers if t["walk"] == w)
            print(f"  walk={w}: {cnt} ({100*cnt/len(losers):.1f}%)")

    print("\n--- B9 verdict (worst 40) ---")
    ge2 = sum(1 for t in worst40 if t["walk"] >= 2)
    print(f"  >=2 walk closes: {ge2}/40 ({100*ge2/40:.0f}%)")

    print("\n--- B10 verdict (longs) ---")
    ll = [t for t in losers if t["side"] == "buy"]
    wl = [t for t in winners if t["side"] == "buy"]
    if ll and wl:
        ml = np.median([t["wick_pct"] for t in ll])
        mw = np.median([t["wick_pct"] for t in wl])
        print(f"  Long losers median wick %: {ml:.1f}")
        print(f"  Long winners median wick %: {mw:.1f}")
        print(f"  Delta (winners - losers): {mw - ml:.1f} pp")

    print("\n--- B10 verdict (shorts) ---")
    ls = [t for t in losers if t["side"] == "sell"]
    ws = [t for t in winners if t["side"] == "sell"]
    if ls and ws:
        ml = np.median([t["wick_pct"] for t in ls])
        mw = np.median([t["wick_pct"] for t in ws])
        print(f"  Short losers median wick %: {ml:.1f}")
        print(f"  Short winners median wick %: {mw:.1f}")
        print(f"  Delta (winners - losers): {mw - ml:.1f} pp")


if __name__ == "__main__":
    main()
