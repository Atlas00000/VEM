#!/usr/bin/env python3
"""Step E: MAE/MFE and hold-time analysis from tester xlsx + M5 bars."""
from __future__ import annotations

import argparse
from datetime import datetime, timedelta, timezone

import numpy as np
import openpyxl

try:
    import MetaTrader5 as mt5
except ImportError:
    mt5 = None

TERMINAL_EXE = r"C:\Program Files\MetaTrader 5\terminal64.exe"
DEFAULT_XLSX = (
    r"C:\Users\emili\AppData\Roaming\MetaQuotes\Terminal"
    r"\D0E8209F77C8CF37AD8BF550E51FF075\MQL5\Profiles\Tester"
    r"\ReportTesterB-23489.xlsx"
)
OUT_MD = (
    r"C:\Users\emili\AppData\Roaming\MetaQuotes\Terminal"
    r"\D0E8209F77C8CF37AD8BF550E51FF075\MQL5\Experts\VEM"
    r"\step-e-results.md"
)

SL_POINTS = 200
POINT = 0.00001  # EURUSD 5-digit
SL_R_PRICE = SL_POINTS * POINT
M5_SEC = 300


def load_trades(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    start = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row[0] == "Time" and row[1] == "Deal":
            start = i + 1
            break
    if start is None:
        wb.close()
        raise RuntimeError("Deals table not found")

    ins: dict[int, tuple] = {}
    for row in ws.iter_rows(min_row=start, values_only=True):
        if row[2] == "EURUSD" and row[4] == "in":
            ins[int(row[1])] = (row[0], row[3], float(row[6]))

    trades = []
    for row in ws.iter_rows(min_row=start, values_only=True):
        if row[2] != "EURUSD" or row[4] != "out":
            continue
        d_in = int(row[1]) - 1
        if d_in not in ins:
            continue
        et, side, entry_px = ins[d_in]
        xt = row[0]
        exit_px = float(row[6])
        if isinstance(et, datetime):
            entry_dt = et if et.tzinfo else et.replace(tzinfo=timezone.utc)
        else:
            entry_dt = datetime.strptime(str(et), "%Y.%m.%d %H:%M:%S").replace(tzinfo=timezone.utc)
        if isinstance(xt, datetime):
            exit_dt = xt if xt.tzinfo else xt.replace(tzinfo=timezone.utc)
        else:
            exit_dt = datetime.strptime(str(xt), "%Y.%m.%d %H:%M:%S").replace(tzinfo=timezone.utc)
        comment = str(row[12] or "").lower()
        if "sl" in comment:
            exit_type = "sl"
        elif "tp" in comment:
            exit_type = "tp"
        else:
            exit_type = "midline"
        trades.append(
            {
                "entry": entry_dt,
                "exit": exit_dt,
                "side": side,
                "entry_px": entry_px,
                "exit_px": exit_px,
                "profit": float(row[10] or 0),
                "exit_type": exit_type,
            }
        )
    wb.close()
    return trades


def load_rates_mt5(t0: datetime, t1: datetime) -> np.ndarray:
    if mt5 is None:
        raise RuntimeError("MetaTrader5 package not installed")
    if not mt5.initialize(path=TERMINAL_EXE):
        raise RuntimeError(f"MT5 init failed: {mt5.last_error()}")
    if not mt5.symbol_select("EURUSD", True):
        raise RuntimeError(f"symbol_select failed: {mt5.last_error()}")

    parts = []
    t = t0
    while t < t1:
        t_end = min(t + timedelta(days=90), t1)
        r = mt5.copy_rates_range("EURUSD", mt5.TIMEFRAME_M5, t, t_end)
        if r is not None and len(r) > 0:
            parts.append(r)
        t = t_end
    mt5.shutdown()
    if not parts:
        raise RuntimeError("No M5 rates from MT5")
    rates = np.concatenate(parts)
    rates = rates[np.argsort(rates["time"])]
    _, ui = np.unique(rates["time"], return_index=True)
    return rates[ui]


def compute_excursions(
    trades: list[dict], rates: np.ndarray, sl_r_price: float
) -> tuple[list[dict], int]:
    t = rates["time"].astype(np.int64)
    h, l = rates["high"], rates["low"]
    skip = 0

    for tr in trades:
        e_ts = int(tr["entry"].timestamp())
        x_ts = int(tr["exit"].timestamp())
        i0 = int(np.searchsorted(t, e_ts, side="left"))
        i1 = int(np.searchsorted(t, x_ts, side="right"))
        if i0 >= len(t) or i1 <= i0:
            skip += 1
            tr["skip"] = True
            continue

        seg_h = h[i0:i1]
        seg_l = l[i0:i1]
        ep = tr["entry_px"]
        is_long = tr["side"] == "buy"

        if is_long:
            mae_px = max(0.0, ep - float(seg_l.min()))
            mfe_px = max(0.0, float(seg_h.max()) - ep)
        else:
            mae_px = max(0.0, float(seg_h.max()) - ep)
            mfe_px = max(0.0, ep - float(seg_l.min()))

        tr["mae_r"] = mae_px / sl_r_price
        tr["mfe_r"] = mfe_px / sl_r_price
        tr["bars_held"] = max(1, (x_ts - e_ts) // M5_SEC)
        tr["skip"] = False

    return trades, skip


def pct(arr, q):
    a = np.array([x for x in arr if x is not None and not np.isnan(x)])
    if len(a) == 0:
        return float("nan")
    return float(np.percentile(a, q))


def summarize_group(rows: list[dict], key: str) -> dict:
    vals = [r[key] for r in rows if not r.get("skip")]
    if not vals:
        return {}
    a = np.array(vals)
    return {
        "n": len(a),
        "median": float(np.median(a)),
        "p75": float(np.percentile(a, 75)),
        "p90": float(np.percentile(a, 90)),
        "mean": float(a.mean()),
    }


def md_table(headers, rows):
    lines = ["| " + " | ".join(headers) + " |", "| " + " | ".join("---" for _ in headers) + " |"]
    for r in rows:
        lines.append("| " + " | ".join(str(c) for c in r) + " |")
    return "\n".join(lines)


def build_report(trades: list[dict], skip: int, xlsx: str, sl_points: int, sl_r_price: float) -> str:
    ok = [t for t in trades if not t.get("skip")]
    wins = [t for t in ok if t["profit"] > 0]
    losses = [t for t in ok if t["profit"] <= 0]

    w_mae = summarize_group(wins, "mae_r")
    l_mfe = summarize_group(losses, "mfe_r")
    w_mfe = summarize_group(wins, "mfe_r")
    l_mae = summarize_group(losses, "mae_r")

    losers_hi_05 = sum(1 for t in losses if t["mfe_r"] > 0.5) / len(losses) * 100 if losses else 0
    losers_hi_08 = sum(1 for t in losses if t["mfe_r"] > 0.8) / len(losses) * 100 if losses else 0
    winners_hi_08_mae = sum(1 for t in wins if t["mae_r"] > 0.8) / len(wins) * 100 if wins else 0

    exit_mix = {}
    for t in ok:
        exit_mix[t["exit_type"]] = exit_mix.get(t["exit_type"], 0) + 1

    w_bars = summarize_group(wins, "bars_held")
    l_bars = summarize_group(losses, "bars_held")

    rec_lines = []
    if w_mae.get("p75", 0) < 0.85:
        rec_lines.append(
            "- **E3 SL:** Winner 75th MAE ~{:.2f}R — SL at 1R rarely threatened on winners; **widening SL is not the first fix**.".format(
                w_mae.get("p75", 0)
            )
        )
    if l_mfe.get("median", 0) > 0.35 and losers_hi_05 > 25:
        rec_lines.append(
            "- **E4 exit:** {:.1f}% of losers reached **>0.5R MFE** (median loser MFE {:.2f}R) — trades often worked then reversed; **midline exit is appropriate**; test **tighter capture** (lower `inp_tp_rr` backup) only after confirming midline share.".format(
                losers_hi_05, l_mfe.get("median", 0)
            )
        )
    elif l_mfe.get("median", 0) < 0.25 and losers_hi_05 < 20:
        rec_lines.append(
            "- **E4 exit:** Losers show **low MFE** (median {:.2f}R, {:.1f}% >0.5R) — failures are fast; **habitat filters (session done) > exit tuning**.".format(
                l_mfe.get("median", 0), losers_hi_05
            )
        )
    else:
        rec_lines.append(
            "- **E4 exit:** Mixed loser MFE (median {:.2f}R, {:.1f}% >0.5R) — prioritize **one** E6 test: slightly lower `inp_tp_rr` (e.g. 1.5→1.2) with midline on, not wider SL.".format(
                l_mfe.get("median", 0), losers_hi_05
            )
        )

    e6 = (
        "**E6 proposal (single change):** Keep `inp_exit_bb_midline=true`. "
        "If losers >0.5R MFE rate ≥25%, test `vem5m_e1_tp_rr_1.2.set` with `inp_tp_rr=1.2` (same session filter). "
        "If losers mostly low MFE, defer exit retest → Step D6 BB width after documenting skip."
    )
    if losers_hi_05 >= 25:
        e6 = (
            "**E6 proposal (single change):** New set `vem5m_e1_tp_rr_12.set` — copy `vem5m_d1_session.set`, "
            "set `inp_tp_rr=1.2` (from 1.5). Midline stays on. Retest IS + OOS vs session control row."
        )

    lines = [
        "# Step E — MAE/MFE analysis",
        "",
        f"**Source:** `{xlsx}` · **SL for 1R:** {sl_points} pts ({sl_r_price:.5f}) · "
        f"**Trades analyzed:** {len(ok)} / {len(trades)} ({skip} skipped — no bar overlap)",
        "",
        "## E1 — Winner MAE (R)",
        "",
        md_table(
            ["Stat", "Winners"],
            [
                ["n", w_mae.get("n", 0)],
                ["Median MAE", f"{w_mae.get('median', 0):.2f}R"],
                ["75th %ile MAE", f"{w_mae.get('p75', 0):.2f}R"],
                ["90th %ile MAE", f"{w_mae.get('p90', 0):.2f}R"],
                ["% winners with MAE > 0.8R", f"{winners_hi_08_mae:.1f}%"],
            ],
        ),
        "",
        f"**vs SL:** 75th percentile MAE {'<' if w_mae.get('p75', 99) < 1.0 else '≥'} 1R — "
        + (
            "winners rarely need full SL room."
            if w_mae.get("p75", 99) < 1.0
            else "some winners draw down near SL before midline exit."
        ),
        "",
        "## E2 — Loser MFE (R)",
        "",
        md_table(
            ["Stat", "Losers"],
            [
                ["n", l_mfe.get("n", 0)],
                ["Median MFE", f"{l_mfe.get('median', 0):.2f}R"],
                ["75th %ile MFE", f"{l_mfe.get('p75', 0):.2f}R"],
                ["% losers MFE > 0.5R", f"{losers_hi_05:.1f}%"],
                ["% losers MFE > 0.8R", f"{losers_hi_08:.1f}%"],
            ],
        ),
        "",
        "## E3–E4 — Decisions",
        "",
        *rec_lines,
        "",
        e6,
        "",
        "## E5 — Hold time (M5 bars)",
        "",
        md_table(
            ["Stat", "Winners", "Losers"],
            [
                ["Median bars", f"{w_bars.get('median', 0):.0f}", f"{l_bars.get('median', 0):.0f}"],
                ["75th %ile", f"{w_bars.get('p75', 0):.0f}", f"{l_bars.get('p75', 0):.0f}"],
            ],
        ),
        "",
        "## Exit type mix (deals comment)",
        "",
        md_table(
            ["Exit", "Count", "%"],
            [
                [k, v, f"{100*v/len(ok):.1f}%"]
                for k, v in sorted(exit_mix.items(), key=lambda x: -x[1])
            ],
        ),
        "",
        "## Loser MAE / Winner MFE (context)",
        "",
        md_table(
            ["Group", "Median MAE (R)", "Median MFE (R)"],
            [
                ["Losers", f"{l_mae.get('median', 0):.2f}", f"{l_mfe.get('median', 0):.2f}"],
                ["Winners", f"{w_mae.get('median', 0):.2f}", f"{w_mfe.get('median', 0):.2f}"],
            ],
        ),
        "",
        "---",
        "",
        "**Checklist:** See `edge-discovery.md` Step E — mark E1–E5 from this file; run **E6** in Strategy Tester after creating proposed `.set`.",
    ]
    return "\n".join(lines)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--out", default=OUT_MD)
    ap.add_argument("--sl-points", type=int, default=SL_POINTS)
    args = ap.parse_args()

    sl_points = args.sl_points
    sl_r_price = sl_points * POINT

    trades = load_trades(args.xlsx)
    if not trades:
        raise SystemExit("No trades loaded")

    t0 = min(t["entry"] for t in trades) - timedelta(days=2)
    t1 = max(t["exit"] for t in trades) + timedelta(days=2)
    rates = load_rates_mt5(t0, t1)
    trades, skip = compute_excursions(trades, rates, sl_r_price)

    report = build_report(trades, skip, args.xlsx, sl_points, sl_r_price)
    with open(args.out, "w", encoding="utf-8") as f:
        f.write(report)
    print(f"Wrote {args.out} ({len(trades)-skip} trades, {skip} skipped)")


if __name__ == "__main__":
    main()
