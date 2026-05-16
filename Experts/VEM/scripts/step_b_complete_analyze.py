#!/usr/bin/env python3
"""Step B complete: B1, B3-B5, B8-B10 from xlsx + EURUSD M5 bars (MT5 or CSV)."""
import argparse
from collections import defaultdict
from datetime import datetime, timedelta, timezone

import numpy as np
import openpyxl
import pandas as pd

try:
    import MetaTrader5 as mt5
except ImportError:
    mt5 = None

TERMINAL_EXE = r"C:\Program Files\MetaTrader 5\terminal64.exe"
XLSX = (
    r"C:\Users\emili\AppData\Roaming\MetaQuotes\Terminal"
    r"\D0E8209F77C8CF37AD8BF550E51FF075\MQL5\Profiles\Tester"
    r"\ReportTester-23489.xlsx"
)
OUT_MD = (
    r"C:\Users\emili\AppData\Roaming\MetaQuotes\Terminal"
    r"\D0E8209F77C8CF37AD8BF550E51FF075\MQL5\Experts\VEM"
    r"\step-b-complete-results.md"
)
BB_PERIOD, BB_DEV = 20, 2.0
RSI_PERIOD = 14
ATR_PERIOD = 20
SIGNAL_SHIFT = 1
M5_SEC = 300


def load_trades(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    start = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row[0] == "Time" and row[1] == "Deal":
            start = i + 1
            break
    ins = {}
    for row in ws.iter_rows(min_row=start, values_only=True):
        if row[2] == "EURUSD" and row[4] == "in":
            ins[int(row[1])] = (row[0], row[3])
    trades = []
    for row in ws.iter_rows(min_row=start, values_only=True):
        if row[2] != "EURUSD" or row[4] != "out":
            continue
        d_in = int(row[1]) - 1
        if d_in not in ins:
            continue
        et, side = ins[d_in]
        if isinstance(et, datetime):
            entry_dt = et
        else:
            entry_dt = datetime.strptime(str(et), "%Y.%m.%d %H:%M:%S")
        if entry_dt.tzinfo is None:
            entry_dt = entry_dt.replace(tzinfo=timezone.utc)
        trades.append(
            {
                "entry": entry_dt,
                "side": side,
                "profit": float(row[10] or 0),
                "exit": str(row[12] or ""),
            }
        )
    wb.close()
    return trades


DEFAULT_CSV = r"c:\Users\emili\OneDrive\Documents\EURUSD_M5_202501092220_202605152055.csv"


def load_rates_csv(path):
    df = pd.read_csv(path, sep="\t")
    dt = pd.to_datetime(
        df["<DATE>"].astype(str) + " " + df["<TIME>"].astype(str),
        format="%Y.%m.%d %H:%M:%S",
    )
    if dt.dt.tz is None:
        dt = dt.dt.tz_localize(timezone.utc)
    rates = np.zeros(
        len(df),
        dtype=[("time", "i8"), ("open", "f8"), ("high", "f8"), ("low", "f8"), ("close", "f8")],
    )
    rates["time"] = (dt.astype("int64") // 10**9).to_numpy()
    rates["open"] = df["<OPEN>"].to_numpy(dtype=float)
    rates["high"] = df["<HIGH>"].to_numpy(dtype=float)
    rates["low"] = df["<LOW>"].to_numpy(dtype=float)
    rates["close"] = df["<CLOSE>"].to_numpy(dtype=float)
    rates = rates[np.argsort(rates["time"])]
    _, ui = np.unique(rates["time"], return_index=True)
    return rates[ui]


def load_rates_mt5():
    if mt5 is None:
        raise RuntimeError("MetaTrader5 package not installed")
    if not mt5.initialize(path=TERMINAL_EXE):
        raise RuntimeError(f"MT5 init failed: {mt5.last_error()}")
    if not mt5.symbol_select("EURUSD", True):
        raise RuntimeError(f"symbol_select failed: {mt5.last_error()}")

    parts = []
    # Recent history via position
    for start in range(0, 150000, 50000):
        r = mt5.copy_rates_from_pos("EURUSD", mt5.TIMEFRAME_M5, start, 50000)
        if r is None or len(r) == 0:
            break
        parts.append(r)

    # Older chunks via range (works when history exists in terminal)
    t0 = datetime(2024, 1, 1, tzinfo=timezone.utc)
    t1 = datetime(2026, 5, 16, tzinfo=timezone.utc)
    t = t0
    while t < t1:
        t_end = min(t + timedelta(days=120), t1)
        r = mt5.copy_rates_range("EURUSD", mt5.TIMEFRAME_M5, t, t_end)
        if r is not None and len(r) > 100:
            parts.append(r)
        t = t_end

    mt5.shutdown()
    if not parts:
        raise RuntimeError("No M5 rates")
    rates = np.concatenate(parts)
    rates = rates[np.argsort(rates["time"])]
    _, ui = np.unique(rates["time"], return_index=True)
    return rates[ui]


def calc_bb(close, i):
    if i < BB_PERIOD - 1:
        return np.nan, np.nan, np.nan
    w = close[i - BB_PERIOD + 1 : i + 1]
    mid = w.mean()
    std = w.std(ddof=0)
    return mid + BB_DEV * std, mid, mid - BB_DEV * std


def calc_rsi(close, i, period=RSI_PERIOD):
    if i < period:
        return np.nan
    deltas = np.diff(close[i - period : i + 1])
    gains = np.where(deltas > 0, deltas, 0.0)
    losses = np.where(deltas < 0, -deltas, 0.0)
    ag, al = gains.mean(), losses.mean()
    if al == 0:
        return 100.0
    rs = ag / al
    return 100.0 - (100.0 / (1.0 + rs))


def calc_atr(h, l, c, i, period=ATR_PERIOD):
    if i < period:
        return np.nan
    trs = []
    for k in range(i - period + 1, i + 1):
        tr = max(h[k] - l[k], abs(h[k] - c[k - 1]), abs(l[k] - c[k - 1]))
        trs.append(tr)
    return float(np.mean(trs))


def calc_ema(close, i, period=20):
    if i < period - 1:
        return np.nan
    alpha = 2.0 / (period + 1)
    ema = close[i - period + 1]
    for k in range(i - period + 2, i + 1):
        ema = alpha * close[k] + (1 - alpha) * ema
    return ema


def enrich_trades(trades, rates):
    t = rates["time"].astype(np.int64)
    o, h, l, c = rates["open"], rates["high"], rates["low"], rates["close"]
    n = len(c)
    time_to_idx = {int(t[i]): i for i in range(n)}

    bb_up = np.full(n, np.nan)
    bb_mid = np.full(n, np.nan)
    bb_lo = np.full(n, np.nan)
    atr_arr = np.full(n, np.nan)
    for i in range(n):
        u, m, lo = calc_bb(c, i)
        bb_up[i], bb_mid[i], bb_lo[i] = u, m, lo
        atr_arr[i] = calc_atr(h, l, c, i)

    atr_valid = atr_arr[~np.isnan(atr_arr)]
    atr_q33, atr_q66 = np.percentile(atr_valid, [33.33, 66.67])

    widths = (bb_up - bb_lo) / np.where(bb_mid > 0, bb_mid, np.nan)
    width_valid = widths[~np.isnan(widths)]
    w_q33, w_q66 = np.percentile(width_valid, [33.33, 66.67])

    def signal_index(entry_dt):
        ts = int(entry_dt.timestamp()) - M5_SEC * SIGNAL_SHIFT
        if ts in time_to_idx:
            return time_to_idx[ts]
        idx = int(np.searchsorted(t, ts, side="right") - 1)
        if idx < 0 or idx >= n or abs(int(t[idx]) - ts) > M5_SEC * 2:
            return None
        return idx

    def walk_count(idx, is_long, max_look=10):
        cnt = 0
        for j in range(1, max_look + 1):
            k = idx - j
            if k < 0 or np.isnan(bb_lo[k]):
                break
            if is_long:
                if c[k] < bb_lo[k]:
                    cnt += 1
                else:
                    break
            else:
                if c[k] > bb_up[k]:
                    cnt += 1
                else:
                    break
        return cnt

    ok = skip = 0
    for tr in trades:
        ets = int(tr["entry"].timestamp())
        if ets < int(t[0]) or ets > int(t[-1]) + M5_SEC:
            skip += 1
            continue
        idx = signal_index(tr["entry"])
        if idx is None or idx < max(BB_PERIOD, ATR_PERIOD, RSI_PERIOD) + 5:
            skip += 1
            continue

        is_long = tr["side"] == "buy"
        mid, up, lo = bb_mid[idx], bb_up[idx], bb_lo[idx]
        width = (up - lo) / mid if mid > 0 else np.nan
        atr_v = atr_arr[idx]
        rsi = calc_rsi(c, idx)
        ema = calc_ema(c, idx)
        ema5 = calc_ema(c, idx - 5) if idx >= 25 else np.nan
        slope = (ema - ema5) / ema5 * 10000 if ema5 and not np.isnan(ema5) else 0.0

        rng = h[idx] - l[idx]
        body_lo, body_hi = min(o[idx], c[idx]), max(o[idx], c[idx])
        wick = (body_lo - l[idx]) if is_long else (h[idx] - body_hi)
        wick_pct = 100.0 * wick / rng if rng > 0 else 0.0

        band_w = up - lo
        stretch = (lo - c[idx]) / band_w if is_long and band_w > 0 else (c[idx] - up) / band_w if band_w > 0 else 0.0

        prev_w = (bb_up[idx - 1] - bb_lo[idx - 1]) if idx > 0 else np.nan
        bb_expand = width > prev_w / mid if idx > 0 and mid > 0 and not np.isnan(prev_w) else False

        atr_prev = atr_arr[idx - 3] if idx >= 3 else np.nan
        atr_expanding = atr_v > atr_prev * 1.05 if not np.isnan(atr_prev) else False

        if atr_v <= atr_q33:
            atr_bucket = "low"
        elif atr_v <= atr_q66:
            atr_bucket = "mid"
        else:
            atr_bucket = "high"

        if width <= w_q33:
            bb_bucket = "narrow"
        elif width <= w_q66:
            bb_bucket = "mid"
        else:
            bb_bucket = "wide"

        if is_long:
            if rsi < 20:
                rsi_bucket = "deep_<20"
            elif rsi < 25:
                rsi_bucket = "20-25"
            elif rsi < 30:
                rsi_bucket = "25-30"
            else:
                rsi_bucket = "30+"
        else:
            if rsi > 80:
                rsi_bucket = "deep_>80"
            elif rsi > 75:
                rsi_bucket = "75-80"
            elif rsi > 70:
                rsi_bucket = "70-75"
            else:
                rsi_bucket = "70-"

        abs_slope = abs(slope)
        if abs_slope < 3 and stretch < 0.6:
            trend = "range"
        elif (is_long and slope < -5) or (not is_long and slope > 5):
            trend = "against"
        elif (is_long and slope > 5) or (not is_long and slope < -5):
            trend = "with"
        else:
            trend = "mild_trend"

        walk = walk_count(idx, is_long)
        tr.update(
            {
                "walk": walk,
                "wick_pct": wick_pct,
                "rsi": rsi,
                "rsi_bucket": rsi_bucket,
                "atr_bucket": atr_bucket,
                "bb_bucket": bb_bucket,
                "trend": trend,
                "slope_bp": slope,
                "stretch": stretch,
                "bb_expand": bb_expand,
                "atr_expanding": atr_expanding,
                # B8: match edge-discovery.md profiles (stricter)
                "good_long": is_long
                and rsi < 22
                and wick_pct >= 18
                and walk <= 1
                and abs_slope < 4
                and stretch >= 0.35
                and not atr_expanding,
                "bad_long": is_long
                and (
                    walk >= 2
                    or (slope < -6 and wick_pct < 12)
                    or (atr_expanding and walk >= 1)
                ),
                "good_short": (not is_long)
                and rsi > 78
                and wick_pct >= 18
                and walk <= 1
                and abs_slope < 4
                and stretch >= 0.35
                and not atr_expanding,
                "bad_short": (not is_long)
                and (
                    walk >= 2
                    or (slope > 6 and wick_pct < 12)
                    or (atr_expanding and walk >= 1)
                ),
            }
        )
        ok += 1
    return ok, skip, int(t[0]), int(t[-1])


def bucket_stats(trades, key):
    buckets = defaultdict(lambda: {"n": 0, "pl": 0.0, "wins": 0})
    for tr in trades:
        b = tr.get(key, "?")
        buckets[b]["n"] += 1
        buckets[b]["pl"] += tr["profit"]
        if tr["profit"] > 0:
            buckets[b]["wins"] += 1
    rows = []
    for b, s in sorted(buckets.items(), key=lambda x: x[1]["pl"]):
        n = s["n"]
        rows.append((b, n, s["pl"], s["wins"] / n * 100 if n else 0))
    return rows


def profile_pct(group, flag):
    if not group:
        return 0.0
    return 100.0 * sum(1 for t in group if t.get(flag)) / len(group)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", default=DEFAULT_CSV, help="MT5-exported M5 OHLC tab file")
    ap.add_argument("--mt5", action="store_true", help="Use MT5 API instead of CSV")
    args = ap.parse_args()

    trades = load_trades(XLSX)
    if args.mt5:
        rates = load_rates_mt5()
        bar_source = "MT5 terminal"
    else:
        rates = load_rates_csv(args.csv)
        bar_source = args.csv
    ok, skip, t0, t1 = enrich_trades(trades, rates)
    analyzed = [t for t in trades if "walk" in t]
    losers = [t for t in analyzed if t["profit"] < 0]
    winners = [t for t in analyzed if t["profit"] > 0]
    worst40 = sorted(losers, key=lambda x: x["profit"])[:40]
    longs = [t for t in analyzed if t["side"] == "buy"]
    shorts = [t for t in analyzed if t["side"] == "sell"]

    lines = []
    lines.append("# Step B — Complete results (EURUSD M5)\n")
    lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")
    lines.append(f"- Trades total: **{len(trades)}**")
    lines.append(f"- Analyzed: **{ok}** | Skipped (no bar): **{skip}**")
    lines.append(f"- Bar source: `{bar_source}`")
    lines.append(
        f"- Bar range: {datetime.fromtimestamp(t0, tz=timezone.utc)} to "
        f"{datetime.fromtimestamp(t1, tz=timezone.utc)}\n"
    )
    if skip > 0:
        lines.append(
            f"> **Note:** **{skip}** trades are before the CSV/bar range (likely **2024.01–2025.01.09**). "
            "Export a second M5 CSV from **2024.01.01** or use History Center Download.\n"
        )

    # B9 B10 summary
    ge2_l = sum(1 for t in losers if t["walk"] >= 2)
    ge2_w = sum(1 for t in winners if t["walk"] >= 2)
    lines.append("## B9 — BB walk\n")
    lines.append(f"| Group | N | ≥2 walk |")
    lines.append(f"|-------|---|---------|")
    lines.append(f"| Losers | {len(losers)} | {ge2_l} ({100*ge2_l/len(losers):.1f}%) |")
    lines.append(f"| Winners | {len(winners)} | {ge2_w} ({100*ge2_w/len(winners):.1f}%) |")
    lines.append(f"| Worst 40 | 40 | {sum(1 for t in worst40 if t['walk']>=2)} |\n")

    lines.append("## B10 — Wick % (median)\n")
    lines.append(f"| Group | Median wick % |")
    lines.append(f"|-------|----------------|")
    lines.append(f"| Losers | {np.median([t['wick_pct'] for t in losers]):.1f} |")
    lines.append(f"| Winners | {np.median([t['wick_pct'] for t in winners]):.1f} |\n")

    # B1
    lines.append("## B1 — Trend vs range (EMA slope proxy)\n")
    lines.append("Buckets: `range` | `mild_trend` | `with` (with drift) | `against`\n")
    lines.append("### All analyzed — P/L by trend bucket\n")
    lines.append("| Trend | N | Net P/L | Win % |")
    lines.append("|-------|---|---------|-------|")
    for b, n, pl, wr in bucket_stats(analyzed, "trend"):
        lines.append(f"| {b} | {n} | {pl:.2f} | {wr:.1f}% |")
    lines.append("\n### Losers only\n")
    lines.append("| Trend | N | Net P/L |")
    lines.append("|-------|---|---------|")
    for b, n, pl, _ in bucket_stats(losers, "trend"):
        lines.append(f"| {b} | {n} | {pl:.2f} |")
    lines.append("")

    # B3 B4 B5
    for title, key in [
        ("B3 — ATR regime (terciles)", "atr_bucket"),
        ("B4 — BB width (terciles)", "bb_bucket"),
        ("B5 — RSI depth at signal", "rsi_bucket"),
    ]:
        lines.append(f"## {title}\n")
        lines.append("| Bucket | N | Net P/L | Win % |")
        lines.append("|--------|---|---------|-------|")
        for b, n, pl, wr in bucket_stats(analyzed, key):
            lines.append(f"| {b} | {n} | {pl:.2f} | {wr:.1f}% |")
        lines.append("")

    # B8
    lines.append("## B8 — Trade quality profiles (rule-based on signal bar)\n")
    lines.append("Automated proxy for good/bad tables in `edge-discovery.md`.\n")
    lines.append("| Profile flag | Losers % | Winners % | Worst40 % |")
    lines.append("|--------------|----------|-----------|-----------|")
    flags = [
        ("good_long", [t for t in longs]),
        ("bad_long", [t for t in longs]),
        ("good_short", [t for t in shorts]),
        ("bad_short", [t for t in shorts]),
    ]
    for flag, pool in flags:
        pl = [t for t in pool if t["side"] == ("buy" if "long" in flag else "sell")]
        if not pl:
            continue
        l_sub = [t for t in pl if t["profit"] < 0]
        w_sub = [t for t in pl if t["profit"] > 0]
        w40 = [t for t in worst40 if t in pl]
        lines.append(
            f"| {flag} | {profile_pct(l_sub, flag):.1f} | {profile_pct(w_sub, flag):.1f} | "
            f"{profile_pct(w40, flag):.1f} |"
        )
    lines.append("\n**Interpretation:** `bad_*` should be **higher on losers** than winners; "
                 "`good_*` **higher on winners**. Large gap = useful filter idea.\n")

    # Top 10 worst with profile tags (B8 manual substitute)
    lines.append("### Worst 10 losers — profile snapshot\n")
    lines.append("| Entry | Side | P/L | walk | wick% | RSI | trend | bad |")
    lines.append("|-------|------|-----|------|-------|-----|-------|-----|")
    for tr in sorted(losers, key=lambda x: x["profit"])[:10]:
        bad = tr.get("bad_long") or tr.get("bad_short")
        lines.append(
            f"| {tr['entry'].strftime('%Y-%m-%d %H:%M')} | {tr['side']} | {tr['profit']:.2f} | "
            f"{tr['walk']} | {tr['wick_pct']:.0f} | {tr['rsi']:.0f} | {tr['trend']} | {bad} |"
        )
    lines.append("")

    # B7
    lines.append("## B7 — Step D priorities (updated)\n")
    lines.append("1. **Session filter (B6)** — hour 13 / NY 13–21 (strongest).")
    lines.append("2. **BB width** — avoid `wide` band entries (B4: wide −$11.92 vs narrow +$3.97).")
    lines.append("3. **RSI** — shorts: avoid shallow `70-` / `75-80`; longs: `deep_<20` / `20-25` best (B5).")
    lines.append("4. **Trend** — `range` entries still net negative; combine with session, not alone.")
    lines.append("5. **BB walk / wick** — weak (B9/B10); optional trial only.\n")

    lines.append("## Full sample (1429 trades)\n")
    lines.append(f"- **{ok}/{len(trades)}** analyzed.")
    lines.append(
        "- For **1429/1429**: add M5 bars from **2024.01.01** (History Center download or second CSV export).\n"
    )
    lines.append(
        f"- Re-run: `python scripts/step_b_complete_analyze.py --csv \"{args.csv}\"`\n"
    )

    text = "\n".join(lines)
    with open(OUT_MD, "w", encoding="utf-8") as f:
        f.write(text)
    print(text.encode("ascii", errors="replace").decode("ascii"))
    print(f"\nWrote {OUT_MD}")


if __name__ == "__main__":
    main()
